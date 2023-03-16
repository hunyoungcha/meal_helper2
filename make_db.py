#기존 엑셀의 데이터 중 원하는 걸 뽑아와 새 db에 저장하는 파일



import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from pprint import pprint

path='dataset/'
file_list=os.listdir(path)

#음식명과 식재료명을 저장해 데이터프레임을 만듬





#재료 딕셔너리,key에 메뉴이름 value에 재료가 저장됨, 매개변수로는 엑셀의 fillna('')가 된 데이터 프레임을 받음
def mat_list(dataframe):
    a=np.array(dataframe[5:])
    mat_list={}
    key=''
    value=[]
    sw=None
    for i in a:
        if '점심' in i[0]:
            sw=1
            key=i[1]
            value.append(i[2].split(',')[0])
        elif '간식' in i[0]:
            sw=0
            key=''
            value=[]
        elif sw==1:
            if i[1]=='':
                value.append(i[2].split(',')[0])
            elif i[1]!='':
                value=','.join(set(value))
                mat_list[key]=value
                key=''
                value=[]
                key=i[1]
                value.append(i[2].split(',')[0])
    return mat_list


def make_dict():
    fin_dict={}
    for file in file_list:
        for sheet in range(2):
            df=pd.read_excel(f"{path+file}", engine = "openpyxl", sheet_name=sheet, usecols='B,C,D')
            df=df.fillna('')
            fin_dict.update(mat_list(df))
            print(f'#########{file}의 {sheet}번째 완성#########')
    print('#########리스트 만들기 완성#########')
    return fin_dict

def to_xls(fin_dict):
    wb=load_workbook('files/database.xlsx')
    ws=wb['database']
    ws.cell(row=1,column=1,value='menu')
    ws.cell(row=1,column=2,value='mat')
    print('######### 리스트 --> 엑셀 #########')
    for row,(key,value) in enumerate(fin_dict.items()):
        ws.cell(row=row+2,column=1,value=key)
        ws.cell(row=row+2,column=2,value=value)
    wb.save('files/database.xlsx')
    print('------성공------')


to_xls(make_dict())